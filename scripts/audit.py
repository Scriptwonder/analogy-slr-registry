#!/usr/bin/env python3
"""Audit 2.0 — recompute every paper-claimed count and PASS/FAIL it against the draft.

Reproducible replacement for the ad-hoc Jul-8 audit (audit_results.txt).
Revision-plan items WS2.14 (reproducible audit script), WS1.10 (freeze corpus,
recompute everything), WS5.27 (consistency QA gate).

Usage:
  python3 scripts/audit.py [Total.xlsx] [--json [audit_v2_results.json]]
          [--tex Sections/AppendixA.tex] [--bib FP.bib FP-extended.bib]

Prints a markdown report to stdout; exits 1 if any check FAILs.
Update the CLAIMS dict below after the corpus freeze — nothing else should
need touching.

WHY THREE INPUTS (registry + appendix tex + bib), NOT JUST Total.xlsx:
  * Total.xlsx (the registry snapshot) holds the PRISMA funnel — per-database
    export sheets, the 'Filtering' ledger, 'Total' (=ATK screening pool),
    'Total_Random' (ATK decisions Y/N/NA), FR-Filtering(+Count)/FPAnalogy
    (full-text decisions Accept/Reject/Uncertain + CrossCheck), FR-Snowball,
    and FullPaperReading(Analogy) (per-paper extraction rows keyed by FP id).
    It does NOT encode the coding dimensions (generation granularity, LLM vs
    non-LLM, multi/single prompt, HITL, automatic/human evaluation, quality
    dimensions). Those memberships exist ONLY as the appendix tables — this is
    itself a release blocker (the promised registry/codebook release needs the
    codes as registry columns; once added, repoint this script at them).
  * Sections/AppendixA.tex — the five coding tables (category memberships).
  * FP.bib / FP-extended.bib — corpus bibkeys, titles, years (trend figure +
    the "approximately half since 2021" claim).

CORPUS / FUNNEL FILTERS CHOSEN (documented per WS2.14; each is a judgement
call, flagged in the report when it bites):
  F1. A registry row is "data" iff any of its first six cells (doi, type,
      year, source, title, abstract) is non-blank — tolerates the junk rows
      with stray content in later columns and rows missing doi+title.
  F2. Corpus := union of cite keys over the five AppendixA coding tables.
      gentner2001analogical-FP40 is in FP.bib and FR-Snowball(Accept) but is
      coded in NO table, so it is NOT counted. ADJUDICATED OUT at the corpus
      freeze (2026-08-22): secondary/book, EC5 — see corpus-freeze.md.
  F3. Round splits by bibkey: '-FPnn' with nn<=33 => round-1 database search
      (reproduces the paper's 31); '-FPnn' with nn>=34 => round-1 snowball /
      manual additions (reproduces the +13; note FP45/46/47 have no
      FR-Snowball row — their provenance rows live only in FPAnalogy /
      FullPaperReading); keys in SECOND_UPDATE_ADDITIONS => round-3
      second-update additions (frozen 2026-08-22, provenance in
      corpus-freeze.md); any other key without FP suffix => round-2 2026
      update (the 21 FP-extended.bib keys that appear in coding tables; the
      remaining FP-extended keys are Webb/Mitchell-debate anchors, not
      corpus).
  F4. "51 automatic evaluations" is approach-level: lexical row + compositional
      row summed, papers in both rows counted twice (matches §4.3 prose).
      Same convention for the 3/24 human split.
  F5. "Approximately half since 2021" passes iff the >=2021 share of the
      corpus (bib years) is within CLAIMS['since2021_band'].
"""
import argparse, json, os, re, sys
from collections import Counter, OrderedDict

# --------------------------------------------------------------------------
# Paper-claimed values (§4 + abstract of the current draft). Update after the
# corpus freeze; the report will show drift until then.
# --------------------------------------------------------------------------
CLAIMS = {
    # PRISMA funnel (funnel-provenance.md, verified 2026-07-11; screened_atk
    # off-by-one RESOLVED 2026-08-22: 'Total' and 'Total_Random' both hold
    # exactly 5,487 unique records (0 duplicate DOIs/titles inside 'Total');
    # the ledger's After=5488 and the Jul-11 "5,488 rows EXACT" both counted
    # the header row / over-count the pool by one. §3.2/PRISMA prose must
    # report 5,487 screened. See corpus-freeze.md.
    "identified_total": 5841,
    "per_db": {"ieee": 2076, "sd": 2565, "acm": 734, "acl": 223,
               "Wiley": 102, "SpringerLink": 67, "springer3": 74},
    "after_dedup": 5680,
    "prescreen_excluded": {"IC1": 151, "EC1": 23, "EC7": 19},
    "screened_atk": 5487,
    "fulltext_reviewed": 270,          # Total_Random Y + NA
    # corpus (FROZEN 2026-08-22: 65 baseline + 18 second-update additions)
    "corpus_total": 83,
    "corpus_round1_db": 31,
    "corpus_round1_snowball": 13,
    "corpus_round2_update": 21,
    "corpus_second_update": 18,
    # generation (§4.2, table:analogy_gen_granularity / table:llm_generation)
    "gen_total": 35, "gen_lexical": 19, "gen_compositional": 17, "gen_both": 1,
    "gen_lexical_llm": 14,             # "19 lexical (14 of which use LLMs)"
    "llm_total": 19, "llm_multi": 12, "llm_single": 7,
    "hitl": 11, "nonllm": 16,          # HITL = 11/19 = 58% (decision 2026-08-22)
    # evaluation (§4.3, table:evaluation_by_granularity / table:evaluation_model)
    "auto_total": 69, "auto_lexical": 42, "auto_compositional": 27,
    "human_lexical": 4, "human_compositional": 26,
    "both_auto_human": 16,
    "no_formal_eval": 1,               # "One paper ... theoretical model" (salu1994)
    "model_total": 20, "model_relational": 6, "model_distributional": 9,
    "model_cognitive": 2, "model_transformation": 2,
    "model_learned": 1,                # new fifth paradigm row (DeepGAR)
    # quality dimensions (table:target_of_analogy_generation)
    "dim_accuracy": 42, "dim_similarity": 23, "dim_validity": 22,
    "dim_novelty": 4, "dim_humanpref": 20,
    # trend claim: share of corpus published >= 2021. At N=81 this is
    # 54/81 = 66.7% — prose must move from "approximately half" to
    # "about two-thirds".
    "since2021_band": (0.60, 0.75),
    # §4.2 HITL sentence: every cite used as a HITL example must be a member
    # of the HITL row of table:llm_generation. cam/cao DECISION (Option 2,
    # 2026-08-22): bhavya2023cam-FP11 is Multi-step only (its filter is
    # automatic scorers, not human) and was REMOVED from this list — a later
    # prose agent must drop it from the §4.2 HITL example sentence;
    # cao2024llm-FP43 is now coded HITL + Multi-step.
    "text_hitl_examples": [
        "sultan2024parallelparc-FP3",
        "bernstein2024nestingdoll", "cao2024llm-FP43", "ju2025toward-FP44",
        "chen2024analogymate", "wang2024know-FP15", "shao2025unlocking-FP42",
    ],
}

# Camera-ready corpus additions, frozen 2026-08-22 (provenance + coding cells
# logged in corpus-freeze.md). Fourth roster bucket ("3-second-update"):
#   - committed 7 (rebuttal/under-retrieval class, class_screening.md Tier A)
#   - 2026 re-run M-set 3 (rerun_results.txt)
#   - snowball 2.0 additions 4 (snowball-screening.md §B)
#   - debate-cluster entrants 2 (snowball-screening.md §C rule)
# Bib entries live in custom.bib (except webb2023emergent / yang2025emergent,
# which live in FP-extended.bib).
SECOND_UPDATE_ADDITIONS = [
    # committed 7
    "ushio2021bert", "sultan2022life", "jacob2023fame",
    "czinczoll2022scientificcreativeanalogiespretrained", "hu2023incontext",
    "fournier2020analogies", "johnson2025analogies",
    # M-set 3 (2026 update re-run)
    "wang2025anascore", "zhang2025annotators", "das2025prototypical",
    # snowball 2.0 four
    "webb2023zeroshot", "ling2022deepgar", "afantenos2026proportional",
    "combs2025tradeoff",
    # debate-cluster entrants
    "webb2023emergent", "yang2025emergent",
    # snowball round-2 find (2026-08-22, see snowball-round2-screening.md)
    "lippolis2026multimodal",
    # full-period title-query find (2026-08-22, admitted by backward-citation
    # criterion: cited by czinczoll2022scan, fournier2020, afantenos2026)
    "rogers2017toomany",
]

REVIEWER_NAMES = set(filter(None, __import__("os").environ.get("REVIEWER_NAMES", "").split(",")))  # names withheld in release; set env var to re-run on raw sheets
DB_SHEETS = ["acl", "ieee", "sd", "Wiley", "acm", "SpringerLink", "springer3"]

def s(v):
    return "" if v is None else str(v).strip()

def cell(row, i):
    """Safe cell access — openpyxl read-only rows are ragged tuples."""
    return s(row[i]) if i < len(row) else ""

def norm(v):
    return re.sub(r"\s+", " ", s(v)).lower()

# --------------------------------------------------------------------------
# Registry (Total.xlsx)
# --------------------------------------------------------------------------
def load_registry(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    reg = {"path": path, "notes": []}

    def rows(name):
        return list(wb[name].iter_rows(values_only=True))

    def data_rows(name):
        # F1: a row counts iff any of the first six cells is non-blank
        return [r for r in rows(name)[1:] if any(s(c) for c in r[:6])]

    # per-DB retrieval counts
    reg["per_db"] = {n: len(data_rows(n)) for n in DB_SHEETS}
    reg["identified_total"] = sum(reg["per_db"].values())

    # 'Filtering' ledger: label -> (remaining, removed)
    ledger = {}
    for r in rows("Filtering"):
        if cell(r, 0):
            ledger[cell(r, 0)] = [cell(r, i) for i in (1, 2, 3)]
    reg["ledger"] = ledger
    def led_int(key, col):
        try:
            return int(float(ledger[key][col]))
        except (KeyError, ValueError, IndexError):
            return None
    # ledger stage counts include the sheet header row (see corpus-freeze.md);
    # the paper reports the header-corrected pool: ledger value - 1.
    reg["after_dedup"] = led_int("De-duplicate", 0) - 1
    reg["prescreen_excluded"] = {k: led_int(k, 1) for k in ("IC1", "EC1", "EC7")}
    reg["ledger_after"] = led_int("After", 0)

    # ATK pool + decisions
    reg["screened_atk"] = len(data_rows("Total"))
    tr = data_rows("Total_Random")
    dec = Counter(cell(r, 12).upper() for r in tr)
    reg["atk_decisions"] = dict(dec)
    reg["fulltext_reviewed"] = dec.get("Y", 0) + dec.get("NA", 0)
    if dec.get(""):
        reg["notes"].append(f"Total_Random: {dec['']} data row(s) with EMPTY Decision")

    # full-text decisions (context, not gating)
    def dcount(name, dcol, ccol):
        c = Counter()
        for r in rows(name)[1:]:
            if cell(r, 0) or cell(r, 1):
                c[(cell(r, dcol), cell(r, ccol))] += 1
        return c
    reg["fr_filtering"] = dcount("FR-Filtering", 12, 13)
    reg["fr_filtering_count"] = dcount("FR-Filtering(Count)", 12, 13)
    reg["fpanalogy"] = dcount("FPAnalogy", 11, 12)
    reg["fpanalogy_accepted"] = sum(
        n for (d, cc), n in reg["fpanalogy"].items()
        if d.lower() == "accept" or cc.lower() == "accept")

    # snowball decisions
    snow = Counter()
    for r in rows("FR-Snowball")[1:]:
        if cell(r, 0):
            snow[(cell(r, 1), cell(r, 7))] += 1
    reg["snowball"] = snow
    reg["snowball_accept_analogy"] = sum(
        n for (d, m), n in snow.items()
        if d.lower() == "accept" and m.lower() != "metaphor")

    # extraction sheet: FP id -> (doi, year, title); flags duplicate ids
    fp_rows, seen = {}, Counter()
    for r in rows("FullPaperReading(Analogy)")[1:]:
        if not cell(r, 0):
            continue
        try:
            fid = int(float(cell(r, 0)))
        except ValueError:
            continue
        seen[fid] += 1
        try:
            year = int(float(cell(r, 3)))
        except ValueError:
            year = None
        fp_rows.setdefault(fid, (cell(r, 1), year, cell(r, 5)))
    reg["fp_rows"] = fp_rows
    dup = [k for k, n in seen.items() if n > 1]
    if dup:
        reg["notes"].append(
            f"FullPaperReading(Analogy): duplicate Paper ID(s) {dup} "
            "(two different papers share an id — disambiguate before release)")

    # reviewer names anywhere (release scrub check)
    hits = Counter()
    for name in wb.sheetnames:
        for r in rows(name):
            for c in r:
                if s(c) in REVIEWER_NAMES:
                    hits[name] += 1
    reg["reviewer_name_cells"] = dict(hits)

    # junk-row diagnostics on the two screening sheets
    for name in ("Total", "Total_Random", "ieee"):
        odd = sum(1 for r in data_rows(name) if not (cell(r, 0) or cell(r, 4)))
        if odd:
            reg["notes"].append(f"{name}: {odd} data row(s) with neither doi nor title")
    return reg

# --------------------------------------------------------------------------
# Appendix coding tables (Sections/AppendixA.tex)
# --------------------------------------------------------------------------
CITE_RE = re.compile(r"\\cite[tp]?\{([^}]*)\}")
BOLD_RE = re.compile(r"\\textbf\{([^}]*)\}")
TYPE_NAMES = {"llm", "non-llm", "automatic", "human"}

def parse_appendix(path):
    """label -> OrderedDict{row_name: [cite keys]}; row_name is
    'Type/Subrow' for three-column tables."""
    tex = open(path).read()
    tex = re.sub(r"(?<!\\)%.*", "", tex)          # strip comments
    tables = {}
    for tm in re.finditer(r"\\begin\{table\*?\}(.*?)\\end\{table\*?\}", tex, re.S):
        body = tm.group(1)
        lab = re.search(r"\\label\{([^}]*)\}", body)
        tab = re.search(r"\\begin\{tabular\}\{[^}]*\}(.*?)\\end\{tabular\}", body, re.S)
        if not (lab and tab):
            continue
        core = tab.group(1)
        core = core.split(r"\midrule", 1)[-1]      # drop header row
        core = core.replace(r"\bottomrule", "")
        rows = OrderedDict()
        for seg in core.split(r"\midrule"):
            segtype = next((b for b in BOLD_RE.findall(seg)
                            if norm(b) in TYPE_NAMES), None)
            for sub in re.split(r"\\cmidrule(?:\([^)]*\))?\{[^}]*\}", seg):
                labels = [re.sub(r"\s+", " ", b).strip()
                          for b in BOLD_RE.findall(sub)
                          if norm(b) not in TYPE_NAMES]
                cites = [k.strip() for grp in CITE_RE.findall(sub)
                         for k in grp.split(",") if k.strip()]
                if not cites:
                    continue
                name = "".join(labels) or segtype or "(unnamed)"
                name = re.sub(r"\s*/\s*", "/", name)
                if segtype and labels:
                    name = f"{segtype}/{name}"
                rows.setdefault(name, [])
                rows[name] += [c for c in cites if c not in rows[name]]
        tables[lab.group(1)] = rows
    return tables

def trow(tables, label, want):
    """Fetch a table row by fuzzy name; hard-fail loudly if the tex changed."""
    rows = tables.get(label, {})
    for name, cites in rows.items():
        if norm(want) in norm(name):
            return set(cites)
    sys.exit(f"FATAL: row '{want}' not found in table '{label}' "
             f"(have: {list(rows)}) — AppendixA.tex layout changed?")

# --------------------------------------------------------------------------
# Bib files
# --------------------------------------------------------------------------
def parse_bib(path):
    out = OrderedDict()
    entry_re = re.compile(r"^@\w+\{([^,\s]+)\s*,", re.M)
    text = open(path).read()
    marks = list(entry_re.finditer(text))
    for i, m in enumerate(marks):
        chunk = text[m.end(): marks[i + 1].start() if i + 1 < len(marks) else len(text)]
        ym = re.search(r"^\s*year\s*=\s*[{\"]?(\d{4})", chunk, re.M | re.I)
        tm2 = re.search(r"^\s*title\s*=\s*[{\"](.+)", chunk, re.M | re.I)
        title = re.sub(r"[{}]", "", tm2.group(1)).rstrip(",\" ").strip() if tm2 else ""
        out[m.group(1)] = {"year": int(ym.group(1)) if ym else None, "title": title}
    return out

def fpnum(key):
    m = re.search(r"-FP(\d+)$", key)
    return int(m.group(1)) if m else None

# --------------------------------------------------------------------------
# Checks / report
# --------------------------------------------------------------------------
class Report:
    def __init__(self):
        self.checks, self.lines = [], []

    def check(self, name, claimed, derived, note=""):
        ok = claimed == derived
        self.checks.append({"name": name, "claimed": claimed, "derived": derived,
                            "status": "PASS" if ok else "FAIL", "note": note})
        return ok

    def warn(self, name, note):
        self.checks.append({"name": name, "claimed": None, "derived": None,
                            "status": "WARN", "note": note})

    def emit(self, line=""):
        self.lines.append(line)
        print(line)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("registry", nargs="?", default="Total.xlsx")
    ap.add_argument("--tex", default=None,
                    help="AppendixA.tex (default: Sections/AppendixA.tex beside registry)")
    ap.add_argument("--bib", nargs="+", default=None,
                    help="bib files, round-1 first, then FP-extended, then any "
                         "lookup-only bibs (default: FP.bib FP-extended.bib "
                         "custom.bib — custom.bib holds the second-update "
                         "additions' entries and is used for year/title lookup "
                         "only, NOT for the anchors count)")
    ap.add_argument("--json", nargs="?", const="audit_v2_results.json", default=None,
                    help="also write machine-readable results to this path")
    args = ap.parse_args()

    root = os.path.dirname(os.path.abspath(args.registry))
    tex = args.tex or os.path.join(root, "Sections", "AppendixA.tex")
    bibs = args.bib or [os.path.join(root, "FP.bib"),
                        os.path.join(root, "FP-extended.bib"),
                        os.path.join(root, "custom.bib")]

    reg = load_registry(args.registry)
    tables = parse_appendix(tex)
    bib_r1 = parse_bib(bibs[0])
    bib_r2 = parse_bib(bibs[1]) if len(bibs) > 1 else OrderedDict()
    bib_lookup = OrderedDict()          # lookup-only (custom.bib etc.)
    for b in bibs[2:]:
        bib_lookup.update(parse_bib(b))
    bib = {**bib_lookup, **bib_r1, **bib_r2}

    # ---- coding sets from the appendix tables --------------------------------
    gen_lex = trow(tables, "table:analogy_gen_granularity", "Lexical-level")
    gen_comp = trow(tables, "table:analogy_gen_granularity", "Compositional-level")
    hitl = trow(tables, "table:llm_generation", "Human-in-the-loop")
    multi = trow(tables, "table:llm_generation", "Multi-step")
    single = trow(tables, "table:llm_generation", "Single Prompt")
    nonllm = trow(tables, "table:llm_generation", "Non-LLM")
    auto_lex = trow(tables, "table:evaluation_by_granularity", "Automatic/Lexical")
    auto_comp = trow(tables, "table:evaluation_by_granularity", "Automatic/Compositional")
    hum_lex = trow(tables, "table:evaluation_by_granularity", "Human/Lexical")
    hum_comp = trow(tables, "table:evaluation_by_granularity", "Human/Compositional")
    dims = {d: trow(tables, "table:target_of_analogy_generation", d)
            for d in ("Accuracy", "Similarity", "Validity", "Novelty", "Human Preference")}
    models = {d: trow(tables, "table:evaluation_model", d)
              for d in ("Relational Graph-based", "Distributional Semantic",
                        "Cognitive/architectural", "Transformation-based",
                        "Learned Relational Representations")}

    all_coded = set().union(gen_lex, gen_comp, hitl, multi, single, nonllm,
                            auto_lex, auto_comp, hum_lex, hum_comp,
                            *dims.values(), *models.values())
    llm = multi | single

    # ---- corpus assembly (filters F2/F3) -------------------------------------
    corpus = sorted(all_coded)
    r1_db = sorted(k for k in corpus if fpnum(k) is not None and fpnum(k) <= 33)
    r1_snow = sorted(k for k in corpus if fpnum(k) is not None and fpnum(k) >= 34)
    cr = sorted(k for k in corpus if k in SECOND_UPDATE_ADDITIONS)
    r2 = sorted(k for k in corpus
                if fpnum(k) is None and k not in SECOND_UPDATE_ADDITIONS)
    fp_uncoded = sorted(set(bib_r1) - all_coded)
    anchors = sorted(set(bib_r2) - all_coded)
    unknown = sorted(k for k in corpus if k not in bib)
    cr_uncoded = sorted(set(SECOND_UPDATE_ADDITIONS) - all_coded)

    def year_of(k):
        y = bib.get(k, {}).get("year")
        if y is None and fpnum(k) in reg["fp_rows"]:
            y = reg["fp_rows"][fpnum(k)][1]
        return y
    years = {k: year_of(k) for k in corpus}
    hist = Counter(y for y in years.values() if y)
    n_dated = sum(hist.values())
    since2021 = sum(n for y, n in hist.items() if y >= 2021)
    share = since2021 / n_dated if n_dated else 0.0

    R = Report()
    C = CLAIMS

    # ---- funnel checks --------------------------------------------------------
    R.check("identified_total", C["identified_total"], reg["identified_total"])
    for db, exp in C["per_db"].items():
        R.check(f"per_db.{db}", exp, reg["per_db"].get(db))
    R.check("after_dedup", C["after_dedup"], reg["after_dedup"])
    for k, exp in C["prescreen_excluded"].items():
        R.check(f"prescreen_excluded.{k}", exp, reg["prescreen_excluded"].get(k))
    R.check("screened_atk (Total sheet rows)", C["screened_atk"], reg["screened_atk"],
            note=f"'Filtering' ledger After={reg['ledger_after']} over-counts "
                 "by one (RESOLVED 2026-08-22): the pool itself is clean — "
                 "5,487 unique records, 0 duplicate dois/titles; prior 5,488 "
                 "counts included the header row. Report 5,487 in §3.2/PRISMA.")
    R.check("fulltext_reviewed (Y+NA)", C["fulltext_reviewed"], reg["fulltext_reviewed"])

    # ---- corpus checks --------------------------------------------------------
    R.check("corpus_total (coded in appendix tables)", C["corpus_total"], len(corpus))
    R.check("corpus_round1_db (FPid<=33)", C["corpus_round1_db"], len(r1_db))
    R.check("corpus_round1_snowball (FPid>=34)", C["corpus_round1_snowball"], len(r1_snow),
            note="FR-Snowball sheet has only "
                 f"{reg['snowball_accept_analogy']} analogy Accepts — FP45/46/47 "
                 "lack snowball provenance rows")
    R.check("corpus_round2_update", C["corpus_round2_update"], len(r2))
    R.check("corpus_second_update (frozen 18)", C["corpus_second_update"], len(cr))
    R.check("second_update_all_coded", [], cr_uncoded,
            note="every SECOND_UPDATE_ADDITIONS key must appear in the coding "
                 "tables" if cr_uncoded else "")
    if fp_uncoded:
        R.warn("fp_keys_uncoded",
               f"uncoded FP.bib keys: {fp_uncoded} — ADJUDICATED OUT at the "
               "2026-08-22 corpus freeze (gentner2001analogical-FP40: "
               "secondary/book, EC5; see corpus-freeze.md). Expected residue, "
               "not an error.")
    if unknown:
        R.warn("table_cites_missing_from_bib", f"{unknown}")

    # ---- generation -----------------------------------------------------------
    R.check("gen_total (lex ∪ comp)", C["gen_total"], len(gen_lex | gen_comp))
    R.check("gen_lexical", C["gen_lexical"], len(gen_lex))
    R.check("gen_compositional", C["gen_compositional"], len(gen_comp))
    R.check("gen_both", C["gen_both"], len(gen_lex & gen_comp),
            note=f"overlap: {sorted(gen_lex & gen_comp)}")
    R.check("gen_lexical_llm ('14 of which use LLMs')",
            C["gen_lexical_llm"], len(gen_lex & llm))
    R.check("llm_total (multi ∪ single)", C["llm_total"], len(llm))
    R.check("llm_multi", C["llm_multi"], len(multi))
    R.check("llm_single", C["llm_single"], len(single))
    R.check("hitl (of LLM)", C["hitl"], len(hitl))
    R.check("hitl_subset_of_llm", [], sorted(hitl - llm))
    R.check("nonllm", C["nonllm"], len(nonllm))
    missing = [k for k in C["text_hitl_examples"] if k not in hitl]
    R.check("text_hitl_examples_in_table_row", [], missing,
            note="§4.2 names these as HITL examples but they are absent from the "
                 "HITL row of table:llm_generation (cam/cao mismatch, revision-plan "
                 "DECISION item 3)" if missing else
                 "cam/cao Option 2 applied: bhavya2023cam-FP11 intentionally NOT "
                 "in this list (prose agent must fix the §4.2 example sentence)")

    # ---- evaluation -----------------------------------------------------------
    R.check("auto_total (lex+comp rows, F4)", C["auto_total"], len(auto_lex) + len(auto_comp))
    R.check("auto_lexical", C["auto_lexical"], len(auto_lex))
    R.check("auto_compositional", C["auto_compositional"], len(auto_comp))
    R.check("human_lexical", C["human_lexical"], len(hum_lex))
    R.check("human_compositional", C["human_compositional"], len(hum_comp))
    both = (auto_lex | auto_comp) & (hum_lex | hum_comp)
    R.check("both_auto_human", C["both_auto_human"], len(both))
    no_eval = set(corpus) - set().union(auto_lex, auto_comp, hum_lex, hum_comp,
                                        *dims.values(), *models.values())
    R.check("no_formal_eval", C["no_formal_eval"], len(no_eval),
            note=f"members: {sorted(no_eval)}")
    R.check("model_total", C["model_total"], sum(len(v) for v in models.values()))
    for key, row in (("model_relational", "Relational Graph-based"),
                     ("model_distributional", "Distributional Semantic"),
                     ("model_cognitive", "Cognitive/architectural"),
                     ("model_transformation", "Transformation-based"),
                     ("model_learned", "Learned Relational Representations")):
        R.check(key, C[key], len(models[row]))
    for key, row in (("dim_accuracy", "Accuracy"), ("dim_similarity", "Similarity"),
                     ("dim_validity", "Validity"), ("dim_novelty", "Novelty"),
                     ("dim_humanpref", "Human Preference")):
        R.check(key, C[key], len(dims[row]))

    # ---- trend ---------------------------------------------------------------
    lo, hi = C["since2021_band"]
    ok = lo <= share <= hi
    R.checks.append({"name": "since2021_share ('about two-thirds')",
                     "claimed": f"[{lo:.2f}, {hi:.2f}]",
                     "derived": f"{share:.3f} ({since2021}/{n_dated})",
                     "status": "PASS" if ok else "FAIL", "note": ""})
    if n_dated < len(corpus):
        R.warn("undated_corpus_papers",
               f"{len(corpus) - n_dated} corpus paper(s) without a year")

    # ---- registry data-quality notes ------------------------------------------
    for note in reg["notes"]:
        R.warn("registry", note)
    if reg["reviewer_name_cells"]:
        R.warn("reviewer_names_present",
               f"cells containing {sorted(REVIEWER_NAMES)} by sheet: "
               f"{reg['reviewer_name_cells']} — scrub before release")

    # ---- report ---------------------------------------------------------------
    R.emit("# Audit 2.0 — registry / tables / claims consistency")
    R.emit()
    R.emit(f"- registry: `{args.registry}`")
    R.emit(f"- coding tables: `{tex}`")
    R.emit(f"- bib: {', '.join('`%s`' % b for b in bibs)}")
    R.emit()
    R.emit("## PASS/FAIL")
    R.emit()
    R.emit("| check | claimed | derived | status | note |")
    R.emit("|---|---|---|---|---|")
    for c in R.checks:
        R.emit("| {name} | {claimed} | {derived} | {status} | {note} |".format(
            **{k: ("" if v is None else v) for k, v in c.items()}))
    n_fail = sum(1 for c in R.checks if c["status"] == "FAIL")
    n_warn = sum(1 for c in R.checks if c["status"] == "WARN")
    R.emit()
    R.emit(f"**{sum(1 for c in R.checks if c['status'] == 'PASS')} PASS, "
           f"{n_fail} FAIL, {n_warn} WARN**")

    R.emit()
    R.emit("## Full-text decision context (registry, non-gating)")
    R.emit()
    def collapse(pairs):
        out = Counter()
        for (d, _), n in pairs.items():
            out[d or "—"] += n
        return dict(out)
    R.emit(f"- ATK decisions (Total_Random): {reg['atk_decisions']}")
    R.emit(f"- FR-Filtering: {collapse(reg['fr_filtering'])}")
    R.emit(f"- FR-Filtering(Count): {collapse(reg['fr_filtering_count'])}")
    R.emit(f"- FPAnalogy accepted pool: {reg['fpanalogy_accepted']} "
           "(paper reports 31 after dropping post-registry: the two non-corpus "
           "rows [FP26 math-questions, Mitchell-2021 review] plus FP47 counted "
           "under the snowball bucket by rule F3)")
    R.emit(f"- FR-Snowball accepts (analogy-scope): {reg['snowball_accept_analogy']}")
    R.emit(f"- debate-cluster anchors in FP-extended.bib (not corpus): {len(anchors)}")

    R.emit()
    R.emit("## Publication-year histogram (trend figure input)")
    R.emit()
    R.emit("| year | n | | year | n |")
    R.emit("|---|---|---|---|---|")
    ys = sorted(hist)
    halfway = (len(ys) + 1) // 2
    for a, b in zip(ys[:halfway], ys[halfway:] + [None] * halfway):
        right = f" {b} | {hist[b]} " if b else "  |  "
        R.emit(f"| {a} | {hist[a]} | |{right}|")
    R.emit()
    R.emit(f"Share published >= 2021: **{since2021}/{n_dated} = {share:.1%}**")

    R.emit()
    R.emit(f"## Corpus roster ({len(corpus)} coded papers — diff basis for additions)")
    R.emit()
    R.emit("| bibkey | round | year | title |")
    R.emit("|---|---|---|---|")
    roster = []
    for k in (r1_db + r1_snow + r2 + cr):
        rnd = ("1-db" if k in r1_db else "1-snowball" if k in r1_snow
               else "2-update" if k in r2 else "3-second-update")
        title = bib.get(k, {}).get("title") or \
            (reg["fp_rows"].get(fpnum(k), ("", None, ""))[2] if fpnum(k) else "")
        roster.append({"bibkey": k, "round": rnd, "year": years[k], "title": title})
        R.emit(f"| {k} | {rnd} | {years[k] or '?'} | {title[:80]} |")

    if args.json:
        payload = {
            "claims": {k: v for k, v in C.items()},
            "checks": R.checks,
            "funnel": {k: reg[k] for k in
                       ("per_db", "identified_total", "after_dedup",
                        "prescreen_excluded", "ledger_after", "screened_atk",
                        "atk_decisions", "fulltext_reviewed",
                        "fpanalogy_accepted", "snowball_accept_analogy")},
            "sets": {"gen_lexical": sorted(gen_lex), "gen_compositional": sorted(gen_comp),
                     "hitl": sorted(hitl), "llm_multi": sorted(multi),
                     "llm_single": sorted(single), "nonllm": sorted(nonllm),
                     "auto_lexical": sorted(auto_lex), "auto_compositional": sorted(auto_comp),
                     "human_lexical": sorted(hum_lex), "human_compositional": sorted(hum_comp),
                     "both_auto_human": sorted(both),
                     "dimensions": {d: sorted(v) for d, v in dims.items()},
                     "model_types": {d: sorted(v) for d, v in models.items()},
                     "fp_uncoded": fp_uncoded, "anchors": anchors,
                     "second_update_additions": SECOND_UPDATE_ADDITIONS},
            "year_histogram": {str(y): hist[y] for y in sorted(hist)},
            "since2021": {"n": since2021, "of": n_dated, "share": share},
            "corpus": roster,
        }
        with open(args.json, "w") as f:
            json.dump(payload, f, indent=2)
        R.emit()
        R.emit(f"(JSON written to {args.json})")

    sys.exit(1 if n_fail else 0)

if __name__ == "__main__":
    main()
